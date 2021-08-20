import os.path
# import os
import sys
import time

from PySide2.QtCore import QThread, Signal, QFile
from PySide2.QtWidgets import QWidget, QApplication, QFileDialog, QTableWidgetItem
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from window import Ui_Form


class CheckExcel(QWidget):
    def __init__(self):
        super(CheckExcel, self).__init__()
        self.__ui = Ui_Form()
        self.__ui.setupUi(self)
        self.my_thread = MyThread()
        self.my_thread.sinOut.connect(self.receive)
        # self.my_thread.start()
        self.my_signal()
        # self.load_excel('')
        self.file_path = r'1.xlsx'

        ink_path = os.path.join(r'C:\ProgramData\Microsoft\Windows\Start Menu\Programs\StartUp', 'Excel数据监控.lnk')
        if not os.path.exists(ink_path):
            exe_path = "Excel数据监控.exe"
            QFile.link(exe_path, ink_path)

    def receive(self, bo):
        if bo:
            self.load_excel(self.file_path)

    def my_signal(self):
        self.__ui.pushButton.clicked.connect(self.open_excel)
        self.__ui.pushButton_3.clicked.connect(self.start_refresh)
        self.__ui.pushButton_2.clicked.connect(self.stop_refresh)

    def open_excel(self):
        file, filetype = QFileDialog.getOpenFileName(self, "Choose file", '', "*.xlsx")
        self.__ui.label.setText('你选择的文件为: {}'.format(os.path.basename(file)))
        self.file_path = file
        self.load_excel(file)
        self.my_thread.start()

    def load_excel(self, filepath):
        # filepath = r'1.xlsx'
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows_heading, cols_heading = [], []
        excel_row, excel_col = 0, 0
        for each_row in range(4, 100):
            item = ws[f'A{each_row}']
            item_value = item.value if item else ''
            if not item_value and not ws[f'A{each_row + 1}'].value:
                excel_row = each_row
                break
            rows_heading.append(item_value)
            print(item_value)
        for each_col in range(3, 100):
            letter = get_column_letter(each_col)
            item = ws[f'{letter}2']
            item_value = item.value if item else ''
            if not item_value and not ws[f'A{each_col + 1}'].value:
                excel_col = each_col
                break
            cols_heading.append(item_value)
        self.__ui.tableWidget.setRowCount(len(rows_heading))
        self.__ui.tableWidget.setColumnCount(len(cols_heading))
        self.__ui.tableWidget.setVerticalHeaderLabels(rows_heading)
        self.__ui.tableWidget.setHorizontalHeaderLabels(cols_heading)
        all_value, single_value = [], []
        for each_row in range(4, excel_row):
            for each_col in range(3, excel_col):
                letter = get_column_letter(each_col)
                single_value.append(ws[f'{letter}{each_row}'].value)
            single_value = [each for each in single_value if each]
            all_value.append(single_value)
            single_value = []
        for row, each_row in enumerate(all_value):
            for col, each in enumerate(each_row):
                self.__ui.tableWidget.setItem(row, col, QTableWidgetItem())
                self.__ui.tableWidget.item(row, col).setText(str(each))

    def start_refresh(self):
        global start, second
        start = True
        second = self.__ui.spinBox.value()
        self.my_thread.start()

    @staticmethod
    def stop_refresh():
        global start
        start = False

    def closeEvent(self, event):
        global start
        start = False
        self.my_thread.quit()


class MyThread(QThread):
    sinOut = Signal(bool)

    def __init__(self):
        super(MyThread, self).__init__()

    def run(self):
        while start:
            print(1)
            time.sleep(second)
            self.sinOut.emit(True)


if __name__ == '__main__':
    start = True
    second = 2
    app = QApplication(sys.argv)
    check = CheckExcel()
    check.show()
    sys.exit(app.exec_())
